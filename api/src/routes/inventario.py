from fastapi import APIRouter, HTTPException, UploadFile, File, status, Query, Body
from ..config.mongodb import items_collection, pedidos_collection, contadores_collection
from ..models.authmodels import Item, InventarioExcelItem
from bson import ObjectId
from pydantic import BaseModel
from typing import List, Literal, Optional  # Keep this import as it's used in the /bulk endpoint
import openpyxl
import io
import os

router = APIRouter()

# Control de logs: solo mostrar en desarrollo
DEBUG_MODE = os.getenv("DEBUG", "false").lower() == "true"
def debug_log(*args, **kwargs):
    """Función para logs de debug - solo muestra en modo DEBUG"""
    if DEBUG_MODE:
        print(*args, **kwargs)

def generar_codigo_automatico():
    """
    Genera un código automático para items siguiendo el formato ITEM-XXXX
    Incrementa el contador y retorna el código formateado.
    Si no existe el contador, lo inicializa directamente en 271.
    """
    # Verificar si el contador existe
    contador_existente = contadores_collection.find_one({"tipo": "items"})
    
    if not contador_existente:
        # Si no existe, crear con secuencia 270 para que el primer incremento resulte en 271
        contadores_collection.insert_one({"tipo": "items", "secuencia": 270})
    
    # Incrementar y obtener el nuevo número
    contador_doc = contadores_collection.find_one_and_update(
        {"tipo": "items"},
        {"$inc": {"secuencia": 1}},
        return_document=True
    )
    numero = contador_doc.get("secuencia", 271)
    return f"ITEM-{numero:04d}"

class ActualizarExistenciaRequest(BaseModel):
    cantidad: float
    tipo: Literal['cargar', 'descargar']
    sucursal: Optional[Literal['sucursal1', 'sucursal2']] = 'sucursal1'  # Por defecto sucursal1

@router.post("/cargar-existencias-desde-pedido")
async def cargar_existencias_desde_pedido(pedido_id: str = Body(..., embed=True)):
    """Carga cantidades al inventario a partir de un pedido específico.
    
    Request Body:
    {
        "pedido_id": "69042b91a9a8ebdaf861c3f0"
    }
    
    Para cada item del pedido:
    - Si existe en inventario: incrementa CANTIDAD con item.cantidad del pedido
    - Si no existe: crea un nuevo item en el inventario con CANTIDAD igual a cantidad del pedido
    
    NOTA: Este endpoint incrementa el campo 'cantidad' en lugar de 'existencia'
    """
    try:
        # Obtener el pedido
        pedido_obj_id = ObjectId(pedido_id)
        pedido = pedidos_collection.find_one({"_id": pedido_obj_id})
        
        if not pedido:
            raise HTTPException(status_code=404, detail="Pedido no encontrado")
        
        items_actualizados = 0
        items_creados = 0
        
        print(f"DEBUG CARGAR EXISTENCIAS: Procesando pedido {pedido_id}")
        
        # Para cada item del pedido
        for idx, item_pedido in enumerate(pedido.get("items", [])):
            codigo_item_raw = item_pedido.get("codigo") or item_pedido.get("id") or item_pedido.get("_id")
            if not codigo_item_raw:
                print(f"DEBUG CARGAR EXISTENCIAS: Item {idx} sin código, saltando")
                continue
            
            # Normalizar código: trim y convertir a string
            codigo_item = str(codigo_item_raw).strip()
            cantidad = int(item_pedido.get("cantidad", 0))
            
            print(f"DEBUG CARGAR EXISTENCIAS: Item {idx} - código: '{codigo_item}', cantidad: {cantidad}")
            
            if cantidad <= 0:
                print(f"DEBUG CARGAR EXISTENCIAS: Item {idx} con cantidad <= 0, saltando")
                continue
            
            # Buscar item en inventario con código normalizado
            # Intentar búsqueda exacta primero
            item_inventario = items_collection.find_one({"codigo": codigo_item})
            codigo_bd = codigo_item  # Código que usaremos para actualizar
            
            # Si no se encuentra, intentar búsqueda sin distinguir mayúsculas/minúsculas
            if not item_inventario:
                print(f"DEBUG CARGAR EXISTENCIAS: No se encontró item con código exacto '{codigo_item}', intentando búsqueda case-insensitive")
                item_inventario = items_collection.find_one({
                    "codigo": {"$regex": f"^{codigo_item}$", "$options": "i"}
                })
                if item_inventario:
                    # Usar el código exacto de la BD, no el normalizado del pedido
                    codigo_bd = item_inventario.get("codigo")
                    print(f"DEBUG CARGAR EXISTENCIAS: Item encontrado con búsqueda flexible. Código en BD: '{codigo_bd}', código pedido: '{codigo_item}'")
            
            if item_inventario:
                codigo_bd_real = item_inventario.get("codigo", codigo_bd)
                item_id = item_inventario.get("_id")
                
                print(f"DEBUG CARGAR EXISTENCIAS: Item encontrado en inventario:")
                print(f"  - _id: {item_id}")
                print(f"  - código en BD: '{codigo_bd_real}'")
                print(f"  - código buscado: '{codigo_item}'")
                print(f"  - cantidad actual: {item_inventario.get('cantidad')}")
                print(f"  - existencia actual: {item_inventario.get('existencia')}")
                
                # Verificar si el campo cantidad existe y es numérico (usamos cantidad, no existencia)
                cantidad_actual = item_inventario.get("cantidad")
                if cantidad_actual is None:
                    print(f"DEBUG CARGAR EXISTENCIAS: Item '{codigo_bd_real}' no tiene campo 'cantidad', creándolo con valor {cantidad}")
                    result = items_collection.update_one(
                        {"_id": item_id},
                        {"$set": {"cantidad": cantidad}}
                    )
                    print(f"DEBUG CARGAR EXISTENCIAS: Resultado $set cantidad - matched: {result.matched_count}, modified: {result.modified_count}")
                else:
                    # Verificar que sea numérico
                    if not isinstance(cantidad_actual, (int, float)):
                        print(f"WARNING CARGAR EXISTENCIAS: Item '{codigo_bd_real}' tiene cantidad no numérica: {cantidad_actual} (tipo: {type(cantidad_actual)}), convirtiendo a número")
                        try:
                            cantidad_actual = float(cantidad_actual)
                            result = items_collection.update_one(
                                {"_id": item_id},
                                {"$set": {"cantidad": cantidad_actual}}
                            )
                            print(f"DEBUG CARGAR EXISTENCIAS: Resultado $set cantidad convertida - matched: {result.matched_count}, modified: {result.modified_count}")
                        except (ValueError, TypeError) as e:
                            print(f"ERROR CARGAR EXISTENCIAS: No se pudo convertir cantidad a número para item '{codigo_bd_real}': {e}, usando 0")
                            cantidad_actual = 0
                            items_collection.update_one(
                                {"_id": item_id},
                                {"$set": {"cantidad": 0}}
                            )
                    
                    # Incrementar cantidad usando _id para asegurar que actualice el documento correcto
                    print(f"DEBUG CARGAR EXISTENCIAS: Incrementando cantidad de item '{codigo_bd_real}' (_id: {item_id}) de {cantidad_actual} a {cantidad_actual + cantidad}")
                    result = items_collection.update_one(
                        {"_id": item_id},
                        {"$inc": {"cantidad": cantidad}}
                    )
                    print(f"DEBUG CARGAR EXISTENCIAS: Resultado update_one - matched: {result.matched_count}, modified: {result.modified_count}")
                    
                    # Verificar el resultado después del update
                    if result.modified_count == 0:
                        print(f"WARNING CARGAR EXISTENCIAS: update_one no modificó ningún documento. Verificar que el _id sea correcto.")
                        # Re-leer el item para ver su estado actual
                        item_actualizado = items_collection.find_one({"_id": item_id})
                        if item_actualizado:
                            print(f"DEBUG CARGAR EXISTENCIAS: Estado actual del item después del update: cantidad={item_actualizado.get('cantidad')}")
                    else:
                        # Verificar que se incrementó correctamente
                        item_actualizado = items_collection.find_one({"_id": item_id})
                        if item_actualizado:
                            print(f"DEBUG CARGAR EXISTENCIAS: ✓ Item actualizado correctamente. Nueva cantidad: {item_actualizado.get('cantidad')}")
                
                items_actualizados += 1
            else:
                # Crear nuevo item en inventario
                print(f"DEBUG CARGAR EXISTENCIAS: Item '{codigo_item}' no existe en inventario, creándolo nuevo")
                nuevo_item = {
                    "codigo": codigo_item,
                    "nombre": item_pedido.get("nombre", item_pedido.get("descripcion", "")),
                    "descripcion": item_pedido.get("descripcion", ""),
                    "cantidad": cantidad,  # Usar cantidad en lugar de existencia
                    "existencia": 0,  # Dejar existencia en 0 por defecto
                    "precio": item_pedido.get("precio", 0),
                    "costo": item_pedido.get("costo", 0),
                    "costoProduccion": item_pedido.get("costoProduccion", 0),
                    "activo": True,
                    "imagenes": item_pedido.get("imagenes", [])
                }
                result = items_collection.insert_one(nuevo_item)
                print(f"DEBUG CARGAR EXISTENCIAS: Item '{codigo_item}' creado con _id: {result.inserted_id}, cantidad inicial: {cantidad}")
                items_creados += 1
        
        print(f"DEBUG CARGAR EXISTENCIAS: Proceso completado - Actualizados: {items_actualizados}, Creados: {items_creados}")
        
        return {
            "message": "Existencias cargadas al inventario correctamente",
            "items_actualizados": items_actualizados,
            "items_creados": items_creados
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al cargar existencias: {str(e)}")

@router.get("/all")
async def get_all_items(sucursal: Optional[str] = Query(None, description="Filtrar por sucursal: 'sucursal1' o 'sucursal2'")):
    """
    Obtener todos los items del inventario.
    Filtra solo items activos con precio > 0.
    Si se especifica sucursal, incluye información de existencia de esa sucursal.
    """
    # Proyección optimizada: solo campos necesarios
    projection = {
        "_id": 1,
        "codigo": 1,
        "nombre": 1,
        "descripcion": 1,
        "categoria": 1,
        "precio": 1,
        "costo": 1,
        "cantidad": 1,
        "existencia": 1,
        "existencia2": 1,
        "activo": 1,
        "imagenes": 1
    }
    
    # OPTIMIZACIÓN: Limitar a 2000 items más recientes
    # Filtrar items activos con precio > 0
    items = list(items_collection.find({
        "activo": True,
        "precio": {"$gt": 0}
    }, projection)
    .sort("_id", -1)  # Ordenar por _id descendente (más recientes primero)
    .limit(2000))
    
    for item in items:
        item["_id"] = str(item["_id"])
        
        # Si se especifica sucursal, agregar campo de existencia de esa sucursal
        if sucursal:
            if sucursal == "sucursal1":
                # Sucursal 1 puede usar "cantidad" o "existencia"
                existencia = item.get("cantidad") or item.get("existencia", 0)
                item["existencia_sucursal"] = existencia
            elif sucursal == "sucursal2":
                # Sucursal 2 usa "existencia2"
                item["existencia_sucursal"] = item.get("existencia2", 0)
    
    return items

@router.get("/id/{item_id}/")
async def get_item(item_id: str):
    """
    Obtener un item por ID o código.
    Intenta primero con ObjectId, si falla busca por código.
    """
    item = None
    
    # Intentar primero con ObjectId
    try:
        item_obj_id = ObjectId(item_id)
        item = items_collection.find_one({"_id": item_obj_id})
    except Exception:
        # Si no es un ObjectId válido, intentar buscar por código
        item = items_collection.find_one({"codigo": item_id})
    
    if not item:
        raise HTTPException(status_code=404, detail=f"Item no encontrado con ID/código: {item_id}")
    
    item["_id"] = str(item["_id"])
    return item

@router.post("/")
async def create_item(item: Item):
    try:
        debug_log(f"DEBUG CREATE ITEM: === INICIO CREACIÓN ===")
        debug_log(f"DEBUG CREATE ITEM: Item recibido - codigo={item.codigo}, nombre={item.nombre}")
        
        # Si no se proporciona código, generar uno automáticamente
        if not item.codigo or (isinstance(item.codigo, str) and item.codigo.strip() == ""):
            item.codigo = generar_codigo_automatico()
            debug_log(f"DEBUG CREATE ITEM: Código generado automáticamente: {item.codigo}")
        
        # Verificar que el código no exista
        existing_item = items_collection.find_one({"codigo": item.codigo})
        if existing_item:
            debug_log(f"DEBUG CREATE ITEM: ❌ Código ya existe: {item.codigo}")
            raise HTTPException(status_code=400, detail="El item con este código ya existe")
        
        # Convertir a diccionario con todos los campos (exclude_unset=False)
        item_dict = item.dict(by_alias=True, exclude_unset=False)
        
        # Limpiar campos None y asegurar valores por defecto
        item_dict_clean = {}
        for key, value in item_dict.items():
            # Excluir _id si existe (para nuevos items)
            if key == "_id" or key == "id":
                continue
            # Incluir el campo si tiene valor o si es un campo requerido
            if value is not None:
                item_dict_clean[key] = value
        
        # Asegurar campos requeridos con valores por defecto
        item_dict_clean.setdefault("codigo", item.codigo)
        item_dict_clean.setdefault("nombre", item.nombre)
        item_dict_clean.setdefault("descripcion", item.descripcion or "")
        item_dict_clean.setdefault("categoria", item.categoria)
        item_dict_clean.setdefault("precio", float(item.precio))
        item_dict_clean.setdefault("costo", float(item.costo))
        item_dict_clean.setdefault("costoProduccion", float(item.costoProduccion) if item.costoProduccion else 0.0)
        item_dict_clean.setdefault("cantidad", int(item.cantidad) if item.cantidad else 0)
        item_dict_clean.setdefault("existencia", int(item.existencia) if hasattr(item, 'existencia') and item.existencia is not None else 0)
        item_dict_clean.setdefault("existencia2", int(item.existencia2) if hasattr(item, 'existencia2') and item.existencia2 is not None else 0)
        item_dict_clean.setdefault("activo", item.activo if item.activo is not None else True)
        item_dict_clean.setdefault("imagenes", item.imagenes if item.imagenes else [])
        
        # Campos opcionales
        if item.departamento:
            item_dict_clean["departamento"] = item.departamento
        if item.marca:
            item_dict_clean["marca"] = item.marca
        if item.modelo:
            item_dict_clean["modelo"] = item.modelo
        
        debug_log(f"DEBUG CREATE ITEM: Datos a insertar:")
        debug_log(f"  - codigo: {item_dict_clean.get('codigo')}")
        debug_log(f"  - nombre: {item_dict_clean.get('nombre')}")
        debug_log(f"  - categoria: {item_dict_clean.get('categoria')}")
        debug_log(f"  - precio: {item_dict_clean.get('precio')}")
        debug_log(f"  - cantidad: {item_dict_clean.get('cantidad')}")
        debug_log(f"  - activo: {item_dict_clean.get('activo')}")
        
        # Insertar en la base de datos
        result = items_collection.insert_one(item_dict_clean)
        
        if not result.inserted_id:
            debug_log(f"DEBUG CREATE ITEM: ❌ ERROR: No se obtuvo inserted_id")
            raise HTTPException(status_code=500, detail="Error al insertar el item en la base de datos")
        
        debug_log(f"DEBUG CREATE ITEM: ✅ Item insertado con _id: {result.inserted_id}")
        
        # Verificar que realmente se guardó
        item_verificado = items_collection.find_one({"_id": result.inserted_id})
        if not item_verificado:
            debug_log(f"DEBUG CREATE ITEM: ❌ ERROR: Item no encontrado después de insertar")
            raise HTTPException(status_code=500, detail="El item no se guardó correctamente en la base de datos")
        
        debug_log(f"DEBUG CREATE ITEM: ✅ Verificación exitosa - Item guardado correctamente")
        debug_log(f"DEBUG CREATE ITEM: === FIN CREACIÓN ===")
        
        return {
            "message": "Item creado correctamente",
            "id": str(result.inserted_id),
            "codigo": item.codigo
        }
    except HTTPException:
        raise
    except Exception as e:
        debug_log(f"DEBUG CREATE ITEM: ❌ EXCEPCIÓN: {str(e)}")
        debug_log(f"DEBUG CREATE ITEM: Tipo: {type(e).__name__}")
        import traceback
        debug_log(f"DEBUG CREATE ITEM: Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error al crear el item: {str(e)}")

@router.patch("/id/{item_id}/")
async def patch_item(item_id: str, update_data: dict = Body(...)):
    """
    Actualizar parcialmente un item del inventario.
    Permite actualizar solo los campos especificados en el body.
    Acepta ObjectId o código como identificador.
    
    Ejemplo:
    {
        "cantidad": 10.5,
        "precio": 150.0
    }
    """
    try:
        # Buscar item por ObjectId o código
        item = None
        item_obj_id = None
        
        try:
            item_obj_id = ObjectId(item_id)
            item = items_collection.find_one({"_id": item_obj_id})
        except Exception:
            # Si no es un ObjectId válido, buscar por código
            item = items_collection.find_one({"codigo": item_id})
            if item:
                item_obj_id = item["_id"]

        if not item:
            raise HTTPException(status_code=404, detail=f"Item no encontrado con ID/código: {item_id}")

        # Filtrar campos que no deben actualizarse
        excluded_fields = {"_id", "id"}
        update_data_clean = {k: v for k, v in update_data.items() if k not in excluded_fields}
        
        # Verificar que hay datos para actualizar
        if not update_data_clean:
            raise HTTPException(status_code=400, detail="No hay datos para actualizar")

        # Realizar la actualización
        result = items_collection.update_one(
            {"_id": item_obj_id},
            {"$set": update_data_clean}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Item no encontrado")
        
        # Obtener el item actualizado
        item_actualizado = items_collection.find_one({"_id": item_obj_id})
        item_actualizado["_id"] = str(item_actualizado["_id"])
            
        return {"message": "Item actualizado correctamente", "id": item_id, "item": item_actualizado}
        
    except HTTPException:
        # Re-lanzar HTTPExceptions (errores conocidos)
        raise
    except Exception as e:
        # Capturar cualquier otro error inesperado
        raise HTTPException(status_code=500, detail=f"Error interno del servidor: {str(e)}")

@router.put("/id/{item_id}/")
async def update_item(item_id: str, item: Item):
    try:
        # Buscar item por ObjectId o código
        existing_item = None
        item_obj_id = None
        
        try:
            item_obj_id = ObjectId(item_id)
            existing_item = items_collection.find_one({"_id": item_obj_id})
        except Exception:
            # Si no es un ObjectId válido, buscar por código
            existing_item = items_collection.find_one({"codigo": item_id})
            if existing_item:
                item_obj_id = existing_item["_id"]
        
        if not existing_item:
            raise HTTPException(status_code=404, detail=f"Item no encontrado con ID/código: {item_id}")

        # Preparar datos para actualización
        update_data = item.dict(exclude_unset=True, by_alias=True, exclude={"_id", "id"})
        
        # Verificar que hay datos para actualizar
        if not update_data:
            raise HTTPException(status_code=400, detail="No hay datos para actualizar")

        # Realizar la actualización
        result = items_collection.update_one(
            {"_id": item_obj_id},
            {"$set": update_data}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Item no encontrado")
            
        return {"message": "Item actualizado correctamente", "id": item_id}
        
    except HTTPException:
        # Re-lanzar HTTPExceptions (errores conocidos)
        raise
    except Exception as e:
        # Capturar cualquier otro error inesperado
        raise HTTPException(status_code=500, detail=f"Error interno del servidor: {str(e)}")

@router.post("/preview-excel")
async def preview_inventory_excel(file: UploadFile = File(...)):
    """Endpoint para previsualizar el contenido del Excel antes de cargarlo"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Formato de archivo no válido. Se espera un archivo Excel (.xlsx o .xls)")

    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = workbook.active

        # Asumiendo que la primera fila son los encabezados
        headers = [cell.value for cell in sheet[1]]
        
        # Validar encabezados requeridos (acepta variaciones)
        headers_lower = {str(h).strip().lower() if h else "" for h in headers}
        if "codigo" not in headers_lower or "descripcion" not in headers_lower or "precio" not in headers_lower or "costo" not in headers_lower:
            raise HTTPException(status_code=400, detail="Faltan encabezados requeridos: codigo, descripcion, precio, costo")

        preview_data = []
        for row_index in range(2, sheet.max_row + 1):
            row_data_raw = {headers[i]: cell.value for i, cell in enumerate(sheet[row_index])}
            
            # Normalizar nombres de columnas para mapear a existencia/existencia2
            row_data = {}
            for key, value in row_data_raw.items():
                key_lower = str(key).strip().lower() if key else ""
                # Mapear variaciones de nombres de columnas
                if key_lower in ["existencia", "sucursal 1", "sucursal1"]:
                    row_data["existencia"] = value
                elif key_lower in ["sucursal 2", "sucursal2", "existencia2"]:
                    row_data["existencia2"] = value
                else:
                    row_data[key] = value
            
            # Solo incluir datos si hay información en la fila
            if any(value is not None for value in row_data.values()):
                preview_data.append(row_data)

        return {
            "total_rows": len(preview_data),
            "headers": headers,
            "data": preview_data
        }

    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al procesar el archivo Excel: {str(e)}")

@router.post("/upload-excel", status_code=status.HTTP_201_CREATED)
async def upload_inventory_excel(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Formato de archivo no válido. Se espera un archivo Excel (.xlsx o .xls)")

    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = workbook.active

        # Asumiendo que la primera fila son los encabezados
        headers = [cell.value for cell in sheet[1]]
        
        # Validar encabezados requeridos
        headers_lower = {str(h).strip().lower() if h else "" for h in headers}
        if "codigo" not in headers_lower or "descripcion" not in headers_lower or "precio" not in headers_lower or "costo" not in headers_lower:
            raise HTTPException(status_code=400, detail="Faltan encabezados requeridos: codigo, descripcion, precio, costo")

        items_to_insert = []
        for row_index in range(2, sheet.max_row + 1):
            row_data_raw = {headers[i]: cell.value for i, cell in enumerate(sheet[row_index])}
            
            # Normalizar nombres de columnas para mapear a existencia/existencia2
            row_data = {}
            for key, value in row_data_raw.items():
                key_lower = str(key).strip().lower() if key else ""
                # Mapear variaciones de nombres de columnas
                if key_lower in ["existencia", "sucursal 1", "sucursal1"]:
                    row_data["existencia"] = value if value is not None else 0
                elif key_lower in ["sucursal 2", "sucursal2", "existencia2"]:
                    row_data["existencia2"] = value if value is not None else 0
                else:
                    row_data[key] = value
            
            # Si no hay existencia mapeada, establecer default
            if "existencia" not in row_data:
                row_data["existencia"] = 0
            if "existencia2" not in row_data:
                row_data["existencia2"] = 0
            
            # Validate with InventarioExcelItem model
            try:
                excel_item = InventarioExcelItem(**row_data)
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Error de validación en la fila {row_index}: {e}")

            # Map to the main Item model
            item_data = Item(
                codigo=excel_item.codigo,
                nombre=excel_item.descripcion, # Defaulting nombre to descripcion
                descripcion=excel_item.descripcion,
                departamento=excel_item.departamento,
                marca=excel_item.marca,
                categoria="General", # Defaulting categoria
                precio=excel_item.precio,
                costo=excel_item.costo,
                # costoProduccion will use its default value from the Item model
                cantidad=0, # Defaulting quantity, as it's not in Excel input
                existencia=excel_item.existencia,
                existencia2=excel_item.existencia2 if excel_item.existencia2 is not None else 0,
                activo=True,
                imagenes=[]
            )
            items_to_insert.append(item_data.dict(by_alias=True, exclude_unset=True))

        if not items_to_insert:
            raise HTTPException(status_code=400, detail="No se encontraron datos válidos para insertar en el archivo Excel.")

        # Insert or update items
        inserted_count = 0
        updated_count = 0
        for item_data in items_to_insert:
            # Check if item with this 'codigo' already exists
            existing_item = items_collection.find_one({"codigo": item_data["codigo"]})
            if existing_item:
                items_collection.update_one(
                    {"_id": existing_item["_id"]},
                    {"$set": item_data}
                )
                updated_count += 1
            else:
                items_collection.insert_one(item_data)
                inserted_count += 1

        return {"message": f"Inventario procesado correctamente. Insertados: {inserted_count}, Actualizados: {updated_count}"}

    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al procesar el archivo Excel: {str(e)}")

@router.get("/search", response_model=List[Item])
async def search_items(
    query: str = Query(..., min_length=1, description="Texto de búsqueda para código, descripción, nombre, departamento o marca"),
    limit: int = Query(10, gt=0, description="Número máximo de resultados a devolver"),
    skip: int = Query(0, ge=0, description="Número de resultados a omitir para paginación")
):
    search_filter = {
        "$or": [
            {"codigo": {"$regex": query, "$options": "i"}},
            {"descripcion": {"$regex": query, "$options": "i"}},
            {"nombre": {"$regex": query, "$options": "i"}},
            {"departamento": {"$regex": query, "$options": "i"}},
            {"marca": {"$regex": query, "$options": "i"}}
        ]
    }
    
    items = list(items_collection.find(search_filter).skip(skip).limit(limit))
    for item in items:
        item["_id"] = str(item["_id"])
    return items

@router.post("/bulk")
async def bulk_upsert_items(items: List[Item]):
    inserted_count = 0
    updated_count = 0
    skipped_items = []
    errors = []

    for item_data in items:
        item_dict = item_data.dict(by_alias=True)
        
        # Ensure _id is not set for new items, let MongoDB generate it
        if "_id" in item_dict:
            del item_dict["_id"]

        # Check if item with same codigo already exists
        existing_item = items_collection.find_one({"codigo": item_data.codigo})

        if existing_item:
            # Update existing item
            try:
                update_fields = {
                    "descripcion": item_dict.get("descripcion"),
                    "modelo": item_dict.get("modelo"),
                    "precio": item_dict.get("precio"),
                    "costo": item_dict.get("costo"),
                    "nombre": item_dict.get("nombre"),
                    "categoria": item_dict.get("categoria"),
                    "costoProduccion": item_dict.get("costoProduccion"),
                    "activo": item_dict.get("activo"),
                    "imagenes": item_dict.get("imagenes"),
                }
                # Remove None values to avoid setting them in MongoDB
                update_fields = {k: v for k, v in update_fields.items() if v is not None}

                update_operation = {"$set": update_fields}

                # Set cantidad if provided
                if "cantidad" in item_dict:
                    update_operation["$set"]["cantidad"] = item_dict["cantidad"]
                
                # Set existencia and existencia2 if provided
                if "existencia" in item_dict:
                    update_operation["$set"]["existencia"] = item_dict["existencia"]
                if "existencia2" in item_dict:
                    update_operation["$set"]["existencia2"] = item_dict["existencia2"]

                items_collection.update_one(
                    {"_id": existing_item["_id"]},
                    update_operation
                )
                updated_count += 1
            except Exception as e:
                errors.append({"item": item_data.dict(by_alias=True), "error": str(e), "action": "update"})
        else:
            # Insert new item
            try:
                items_collection.insert_one(item_dict)
                inserted_count += 1
            except Exception as e:
                errors.append({"item": item_data.dict(by_alias=True), "error": str(e), "action": "insert"})

    return {
        "message": f"Procesamiento de carga masiva completado. {inserted_count} items insertados, {updated_count} items actualizados, {len(errors)} con errores.",
        "inserted_count": inserted_count,
        "updated_count": updated_count,
        "errors": errors
    }

@router.post("/{item_id}/existencia")
async def actualizar_existencia(item_id: str, request: ActualizarExistenciaRequest):
    """
    Cargar o descargar existencia de un item en la sucursal especificada
    Body esperado:
    {
        "cantidad": 10.0,
        "tipo": "cargar" o "descargar",
        "sucursal": "sucursal1" o "sucursal2" (opcional, por defecto "sucursal1")
    }
    """
    try:
        # Validar ObjectId
        try:
            item_obj_id = ObjectId(item_id)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"item_id no es un ObjectId válido: {str(e)}")
        
        # Buscar el item
        item = items_collection.find_one({"_id": item_obj_id})
        if not item:
            raise HTTPException(status_code=404, detail="Item no encontrado")
        
        # Determinar qué campo de existencia actualizar según la sucursal
        # Sucursal 1 usa el campo "cantidad", Sucursal 2 usa "existencia2"
        campo_existencia = "cantidad" if request.sucursal == "sucursal1" else "existencia2"
        cantidad_actual = item.get(campo_existencia, 0.0)
        
        # Calcular nueva cantidad
        if request.tipo == "cargar":
            nueva_cantidad = cantidad_actual + request.cantidad
        elif request.tipo == "descargar":
            nueva_cantidad = cantidad_actual - request.cantidad
            # Validar que no sea negativa
            if nueva_cantidad < 0:
                raise HTTPException(
                    status_code=400, 
                    detail=f"No se puede descargar más de lo disponible. Existencia actual en {request.sucursal}: {cantidad_actual}"
                )
        else:
            raise HTTPException(status_code=400, detail="Tipo debe ser 'cargar' o 'descargar'")
        
        # Actualizar la existencia según la sucursal
        result = items_collection.update_one(
            {"_id": item_obj_id},
            {"$set": {campo_existencia: nueva_cantidad}}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Item no encontrado")
        
        # Obtener el item actualizado
        item_actualizado = items_collection.find_one({"_id": item_obj_id})
        
        return {
            "message": f"Existencia {request.tipo}da exitosamente en {request.sucursal}",
            "cantidad_actualizada": item_actualizado.get(campo_existencia, nueva_cantidad),
            "cantidad_anterior": cantidad_actual,
            "cantidad_operacion": request.cantidad,
            "tipo": request.tipo,
            "sucursal": request.sucursal,
            "cantidad": item_actualizado.get("cantidad", 0),  # Sucursal 1
            "existencia": item_actualizado.get("existencia", 0),
            "existencia2": item_actualizado.get("existencia2", 0)  # Sucursal 2
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error interno del servidor: {str(e)}")
